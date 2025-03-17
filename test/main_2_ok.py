import os
import yaml
import magic
import time
import logging
import threading
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any, Iterator
from concurrent.futures import ThreadPoolExecutor, as_completed
from markitdown import MarkItDown
from dataclasses import dataclass, field
import re
from pdfminer.high_level import extract_text as pdf_extract_text
import pandas as pd
import openpyxl
import csv
import json
from datetime import datetime
import argparse
import sys
import win32com.client

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('sensitive_detector.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class ProcessingResult:
    """数据类，用于存储文件处理结果"""
    file_path: str
    mime_type: str
    content: Dict[str, Any]
    sensitive_words: List[Tuple[str, List[int]]] = field(default_factory=list)
    error: Optional[str] = None
    processing_time: float = 0.0

class FileTypeDetector:
    """文件类型检测器，使用 python-magic 库"""
    
    MIME_TYPES = {
        ".txt": "text/plain", ".csv": "text/csv", ".xml": "text/xml", ".html": "text/html",
        ".htm": "text/html", ".json": "application/json", ".yaml": "application/yaml",
        ".yml": "application/yaml", ".md": "text/markdown", ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xls": "application/vnd.ms-excel",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".ppt": "application/vnd.ms-powerpoint",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".zip": "application/zip", ".rar": "application/x-rar", ".7z": "application/x-7z-compressed",
        ".tar": "application/x-tar", ".gz": "application/gzip", ".bz2": "application/x-bzip2",
        ".pdf": "application/pdf", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
        ".gif": "image/gif", ".bmp": "image/bmp", ".svg": "image/svg+xml", ".webp": "image/webp",
        ".mp3": "audio/mpeg", ".mp4": "video/mp4", ".avi": "video/x-msvideo", ".mov": "video/quicktime",
        ".wav": "audio/wav", ".bin": "application/octet-stream", ".exe": "application/x-msdownload",
        ".dll": "application/x-msdownload",
    }
    MIME_TO_EXT = {mime: ext for ext, mime in MIME_TYPES.items()}

    def __init__(self):
        """初始化文件类型检测器"""
        self.mime = magic.Magic(mime=True)
    
    def get_all_files(self, directory: str) -> List[str]:
        """递归获取目录中的所有文件"""
        return [str(f) for f in Path(directory).rglob('*') if f.is_file()]
    
    def get_file_info(self, file_path: str) -> Dict:
        """获取文件信息，包括 MIME 类型和文件头"""
        try:
            mime_type = self.mime.from_file(file_path)
            with open(file_path, "rb") as f:
                file_header = f.read(16).hex().upper()
            file_extension = Path(file_path).suffix.lower()
            return {
                "file_path": file_path,
                "mime_type": mime_type,
                "file_extension": file_extension,
                "file_header": file_header,
                "size": os.path.getsize(file_path)
            }
        except Exception as e:
            logger.error(f"获取文件信息失败: {e} - {file_path}")
            return {"file_path": file_path, "mime_type": "unknown", "error": str(e)}
    
    def detect_file_type(self, file_path: str) -> str:
        try:
            normalized_path = os.path.normpath(file_path.encode('utf-8', errors='replace').decode('utf-8'))
            mime_type = self.mime.from_file(normalized_path)
            return mime_type or self.MIME_TYPES.get(Path(file_path).suffix.lower(), "application/octet-stream")
        except Exception as e:
            logger.error(f"文件类型检测失败 {file_path}: {e}")
            return self.MIME_TYPES.get(Path(file_path).suffix.lower(), "application/octet-stream")

class ContentExtractor:
    """优化后的文件内容提取器"""
    
    def __init__(self):
        """初始化内容提取器"""
        self.md = MarkItDown()
        self.fallback_extractors = {
            'application/msword': self._extract_doc_content,
            'application/vnd.ms-powerpoint': self._extract_ppt_content,
            'application/vnd.ms-excel': self._extract_xls_content,
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document': self._extract_docx_content,
            'application/vnd.openxmlformats-officedocument.presentationml.presentation': self._extract_pptx_content,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': self._extract_xlsx_content,
            'text/csv': self._read_csv_file,
            'application/pdf': self._extract_pdf_content,
        }
    
    def _extract_pdf_content(self, file_path: str) -> Dict[str, Any]:
        print('pdf')
        """使用 PDFMiner 提取 PDF 内容"""
        try:
            text = pdf_extract_text(file_path)
            return {
                'content': text,
                'metadata': {'file_type': 'pdf'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'pdf'},
                'error': str(e)
            }
    
    def _extract_xlsx_content(self, file_path: str) -> Dict[str, Any]:
        """优化 Excel 文件读取，使用 openpyxl 流式读取"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append([str(cell) if cell is not None else '' for cell in row])
                if data:
                    df = pd.DataFrame(data[1:], columns=data[0])
                    sheets[sheet_name] = df
            return self._process_excel_content(sheets)
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'xlsx'},
                'error': f"openpyxl 处理失败: {str(e)}"
            }
    
    def _process_excel_content(self, sheets: Dict) -> Dict[str, Any]:
        print('excel')
        """处理 Excel 内容的辅助方法"""
        content_parts = []
        metadata = {
            'sheets': list(sheets.keys()),
            'total_sheets': len(sheets),
            'row_counts': {}
        }
        for sheet_name, df in sheets.items():
            content_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
            content_parts.append(df.to_string(index=False))
            metadata['row_counts'][sheet_name] = len(df)
        return {
            'content': '\n'.join(content_parts),
            'metadata': {'file_type': 'excel', **metadata},
            'error': None
        }
    


    def _extract_doc_content(self, file_path: str) -> Dict[str, Any]:
        """提取 DOC 文件内容（使用 win32com）"""
        import pythoncom
        try:
            pythoncom.CoInitialize()  # 初始化 COM
            import win32com.client
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(os.path.abspath(file_path))
            content = doc.Range().Text
            doc.Close(False)
            word.Quit()
            return {
                'content': content,
                'metadata': {'file_type': 'doc'},
                'error': None
            }
        except ImportError:
            return {
                'content': '',
                'metadata': {'file_type': 'doc'},
                'error': '缺少 pywin32 依赖，请安装: pip install pywin32'
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'doc'},
                'error': f"win32com 处理失败: {str(e)}"
            }
        finally:
            pythoncom.CoUninitialize()  # 清理 COM
    
    def _extract_docx_content(self, file_path: str) -> Dict[str, Any]:
        print('docx')
        """提取 DOCX 文件内容"""
        try:
            result = self.md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': 'docx'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'docx'},
                'error': str(e)
            }
    
    def _extract_ppt_content(self, file_path: str) -> Dict[str, Any]:
        """提取 PPT 文件内容（依赖 Windows 和 Microsoft PowerPoint）"""
        print('Processing PPT file...')
        try:
            # 使用 pywin32 调用 PowerPoint 提取文本
            ppt = win32com.client.Dispatch("PowerPoint.Application")
            ppt.Visible = 0  # 后台运行
            presentation = ppt.Presentations.Open(file_path)
            text_content = []
            
            # 遍历幻灯片和形状提取文本
            for slide in presentation.Slides:
                for shape in slide.Shapes:
                    if shape.HasTextFrame:
                        text = shape.TextFrame.TextRange.Text
                        text_content.append(text.strip())
            
            content = "\n".join(text_content)
            presentation.Close()
            ppt.Quit()
            
            return {
                'content': content,
                'metadata': {'file_type': 'ppt'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'ppt'},
                'error': str(e)
            }
        finally:
            # 确保 PowerPoint 进程关闭
            if 'presentation' in locals():
                presentation.Close()
            if 'ppt' in locals():
                ppt.Quit()
    
    def _extract_pptx_content(self, file_path: str) -> Dict[str, Any]:
        print('pptx')
        """提取 PPTX 文件内容"""
        try:
            result = self.md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': 'pptx'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'pptx'},
                'error': str(e)
            }
    
    def _extract_xlsx_content(self, file_path: str) -> Dict[str, Any]:
            """优化 Excel 文件读取，使用 openpyxl 流式读取"""
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheets = {}
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append([str(cell) if cell is not None else '' for cell in row])
                    if data:
                        df = pd.DataFrame(data[1:], columns=data[0])
                        sheets[sheet_name] = df
                return self._process_excel_content(sheets)
            except openpyxl.utils.exceptions.InvalidFileException as e:
                return {
                    'content': '',
                    'metadata': {'file_type': 'xlsx'},
                    'error': f"openpyxl 处理失败: 文件格式不受支持 ({str(e)})，请用 Excel 检查文件"
                }
            except Exception as e:
                return {
                    'content': '',
                    'metadata': {'file_type': 'xlsx'},
                    'error': f"openpyxl 处理失败: {str(e)}"
                }
    def _extract_xls_content(self, file_path: str) -> Dict[str, Any]:
        print('xlsx')
        """提取 XLS 文件内容"""
        try:
            sheets = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
            return self._process_excel_content(sheets)
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'xls'},
                'error': str(e)
            }

    
    def _read_csv_file(self, file_path: str) -> Dict[str, Any]:
        print('csv')
        """读取 CSV 文件内容"""
        try:
            df = pd.read_csv(file_path, sep=None, engine='python')
            return {
                'content': df.to_string(index=False),
                'metadata': {'file_type': 'csv'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'csv'},
                'error': str(e)
            }
    
    def extract_content(self, file_path: str, mime_type: str) -> Dict[str, Any]:
        """提取文件内容，根据文件类型选择合适的提取器"""
        if not Path(file_path).exists():
            return {'error': f'文件不存在: {file_path}', 'content': '', 'metadata': {}}
        
        if Path(file_path).stat().st_size == 0:
            return {
                'content': '',
                'metadata': {'file_type': mime_type},
                'error': None,
                'is_empty': True
            }
        
        # 显式检查 .doc 文件
        if mime_type == 'application/msword' or file_path.lower().endswith('.doc'):
            return self._extract_doc_content(file_path)
        
        # 显式检查 .xlsx 文件
        if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or file_path.lower().endswith('.xlsx'):
            return self._extract_xlsx_content(file_path)
        
        if mime_type in self.fallback_extractors:
            return self.fallback_extractors[mime_type](file_path)
        
        try:
            result = self.md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': mime_type, 'converter': 'markitdown'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': mime_type},
                'error': f"MarkItDown 转换失败: {str(e)}"
            }


class SensitiveChecker:
    """敏感内容检查器，使用正则表达式替代 Aho-Corasick"""
    
    def __init__(self, config_path: str = "sensitive_config.yaml"):
        """初始化敏感词配置"""
        self.config = self._load_config(config_path)
        # 合并所有纯字符串关键词为一个正则表达式模式
        self.all_keywords = self.config.get('security_marks', []) + \
                           [kw for cat in self.config.get('sensitive_patterns', {}).values() 
                            for kw in cat.get('keywords', [])]
        # 转义特殊字符并构建正则表达式
        escaped_keywords = [re.escape(kw) for kw in self.all_keywords]
        self.keyword_pattern = re.compile('|'.join(escaped_keywords))
    
    def _load_config(self, config_path: str) -> Dict:
        """加载敏感词配置文件"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except Exception as e:
            logger.error(f"加载敏感词配置失败: {e}")
            return {}
    
    def check_content(self, text: str) -> List[Tuple[str, List[int]]]:
        """检查文本中的敏感词"""
        # 使用正则表达式一次性匹配所有关键词
        keyword_matches = {}
        for match in self.keyword_pattern.finditer(text):
            keyword = match.group()
            if keyword not in keyword_matches:
                keyword_matches[keyword] = []
            keyword_matches[keyword].append(match.start())
        keyword_results = list(keyword_matches.items())
        
        # 处理结构化模式
        structured_results = []
        for pattern, weight in self.config.get('structured_patterns', {}).items():
            matches = list(re.finditer(pattern, text))
            if matches:
                positions = [m.start() for m in matches]
                structured_results.append((pattern, positions))
        
        # 处理数字模式
        number_results = []
        for pattern in self.config.get('number_patterns', []):
            matches = list(re.finditer(pattern, text))
            if matches:
                positions = [m.start() for m in matches]
                number_results.append((pattern, positions))
        
        return keyword_results + structured_results + number_results

class ResultExporter:
    """处理结果导出器"""
    
    def export_to_json(self, results: List[ProcessingResult], output_path: str):
        """导出结果到 JSON 文件"""
        export_data = [
            {
                'file_path': r.file_path,
                'mime_type': r.mime_type,
                'content_preview': r.content.get('content', '')[:200],
                'sensitive_words': [{'word': w, 'positions': p} for w, p in r.sensitive_words],
                'error': r.error,
                'processing_time': r.processing_time
            } for r in results
        ]
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    def export_to_excel(self, results: List[ProcessingResult], output_path: str):
        """导出结果到 Excel 文件"""
        data = [
            {
                '文件路径': r.file_path,
                '文件类型': r.mime_type,
                '敏感词统计': '; '.join([f"{w}({len(p)}次)" for w, p in r.sensitive_words]),
                '处理时间(秒)': round(r.processing_time, 3),
                '错误信息': r.error or ''
            } for r in results
        ]
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine='openpyxl')

class ResultMonitor:
    """结果监控器，用于实时输出处理进度和结果"""
    
    def __init__(self, output_csv: str = "processing_results.csv"):
        self.output_csv = output_csv
        self._init_csv()
    
    def _init_csv(self):
        """初始化 CSV 文件"""
        with open(self.output_csv, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                '处理时间', '文件路径', '文件类型', '文件大小(bytes)', '是否为空',
                '敏感词数量', '敏感词列表', '处理时长(秒)', '状态', '错误信息'
            ])
    
    def record_result(self, result: ProcessingResult):
        """记录单个处理结果"""
        try:
            file_size = Path(result.file_path).stat().st_size
        except:
            file_size = 0
        
        sensitive_words_count = len(result.sensitive_words)
        sensitive_words_list = '; '.join([f"{w}({len(p)}处)" for w, p in result.sensitive_words])
        
        status = '失败' if result.error else ('空文件' if result.content.get('is_empty') else 
                                              '已跳过' if result.content.get('skipped') else '成功')
        
        with open(self.output_csv, 'a', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                result.file_path,
                result.mime_type,
                file_size,
                'Yes' if result.content.get('is_empty') else 'No',
                sensitive_words_count,
                sensitive_words_list,
                round(result.processing_time, 3),
                status,
                result.error or ''
            ])
        
        print(f"[{datetime.now().strftime('%H:%M:%S')}] 处理文件: {result.file_path}")
        print(f"  类型: {result.mime_type}")
        print(f"  大小: {file_size} bytes")
        print(f"  状态: {status}")
        if sensitive_words_count > 0:
            print(f"  发现敏感词: {sensitive_words_list}")
        if result.error:
            print(f"  错误: {result.error}")
        print("-" * 80)

class FileProcessor:
    """优化后的文件处理器主类"""
    
    def __init__(self, config_path: str = "sensitive_config.yaml", 
                 monitor_output: str = "processing_results.csv",
                 chunk_size: int = 1000,
                 max_workers: Optional[int] = None):
        self.detector = FileTypeDetector()
        self.extractor = ContentExtractor()
        self.checker = SensitiveChecker(config_path)
        self.exporter = ResultExporter()
        self.monitor = ResultMonitor(monitor_output)
        self.chunk_size = chunk_size
        self.max_workers = max_workers or (os.cpu_count() or 1) * 2
        self._mime_cache = {}
        self._file_size_cache = {}
    
    def _scan_directory(self, directory: str) -> Iterator[List[str]]:
        """使用生成器分批扫描目录文件"""
        current_chunk = []
        for entry in os.scandir(directory):
            try:
                if entry.is_file(follow_symlinks=False):
                    current_chunk.append(entry.path)
                    if len(current_chunk) >= self.chunk_size:
                        yield current_chunk
                        current_chunk = []
                elif entry.is_dir(follow_symlinks=False):
                    for sub_chunk in self._scan_directory(entry.path):
                        yield sub_chunk
            except (PermissionError, OSError) as e:
                logger.warning(f"访问文件/目录出错 {entry.path}: {e}")
                continue
        if current_chunk:
            yield current_chunk
    
    def _preload_file_info(self, file_paths: List[str]):
        """预加载文件信息以减少 I/O 操作"""
        for file_path in file_paths:
            try:
                stat = os.stat(file_path)
                self._file_size_cache[file_path] = stat.st_size
                mime_type = self.detector.detect_file_type(file_path)
                self._mime_cache[file_path] = mime_type
            except Exception as e:
                logger.warning(f"预加载文件信息失败 {file_path}: {e}")
    
    def _process_file_batch(self, file_paths: List[str]) -> List[ProcessingResult]:
        """处理一批文件"""
        results = []
        self._preload_file_info(file_paths)
        
        current_workers = min(len(file_paths), self.max_workers)
        with ThreadPoolExecutor(max_workers=current_workers) as executor:
            future_to_file = {executor.submit(self.process_file, fp): fp for fp in file_paths}
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    results.append(result)
                    self.monitor.record_result(result)
                except Exception as e:
                    logger.error(f"处理文件异常 {file_path}: {e}")
        
        if len(self._mime_cache) > 10000:
            self._mime_cache.clear()
            self._file_size_cache.clear()
        
        return results
    
    def process_directory(self, directory: str) -> List[ProcessingResult]:
        """处理目录下的所有文件"""
        results = []
        total_files = sum(1 for _ in Path(directory).rglob('*') if _.is_file())
        completed = 0
        
        print(f"\n开始处理目录: {directory}")
        print(f"共发现 {total_files} 个文件")
        print("=" * 80)
        
        try:
            for file_chunk in self._scan_directory(directory):
                chunk_results = self._process_file_batch(file_chunk)
                results.extend(chunk_results)
                completed += len(file_chunk)
                print(f"\n已完成: {completed}/{total_files} ({completed/total_files*100:.1f}%)")
        except KeyboardInterrupt:
            print("\n用户中断处理...")
        except Exception as e:
            logger.error(f"处理目录时发生错误: {e}")
        finally:
            if results:
                try:
                    self.exporter.export_to_json(results, "interrupted_results.json")
                    self.exporter.export_to_excel(results, "interrupted_results.xlsx")
                except Exception as e:
                    logger.error(f"导出中断结果失败: {e}")
        
        print(f"\n处理完成！共处理 {len(results)}/{total_files} 个文件")
        print(f"详细结果已保存至: {self.monitor.output_csv}")
        print("=" * 80)
        return results
    
    def process_file(self, file_path: str) -> ProcessingResult:
        start_time = time.time()
        try:
            mime_type = self._mime_cache.get(file_path) or self.detector.detect_file_type(file_path)
            logger.info(f"Processing file: {file_path} (type: {mime_type})")
            
            file_size = self._file_size_cache.get(file_path, 0)
            if file_size == 0:
                return ProcessingResult(
                    file_path=file_path,
                    mime_type=mime_type,
                    content={'is_empty': True},
                    sensitive_words=[],
                    error=None,
                    processing_time=time.time() - start_time
                )
            
            content = self.extractor.extract_content(file_path, mime_type)
            if content.get('error'):
                logger.warning(f"提取内容失败: {file_path} - MIME: {mime_type} - 错误: {content.get('error')}")
                return ProcessingResult(
                    file_path=file_path,
                    mime_type=mime_type,
                    content=content,
                    sensitive_words=[],
                    error=content.get('error'),
                    processing_time=time.time() - start_time
                )
            
            sensitive_words = self.checker.check_content(content.get('content', ''))
            if sensitive_words:
                logger.info(f"Found {len(sensitive_words)} sensitive words in {file_path}")
            
            return ProcessingResult(
                file_path=file_path,
                mime_type=mime_type,
                content=content,
                sensitive_words=sensitive_words,
                error=None,
                processing_time=time.time() - start_time
            )
        except Exception as e:
            error_msg = f"处理文件失败: {str(e)}"
            logger.error(f"{error_msg} - {file_path}")
            return ProcessingResult(
                file_path=file_path,
                mime_type=mime_type if 'mime_type' in locals() else "unknown",
                content={'error': error_msg},
                processing_time=time.time() - start_time
            )

def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='文件敏感内容检测工具')
    parser.add_argument('path', help='要处理的文件或目录路径')
    parser.add_argument('--config', default='sensitive_config.yaml', help='敏感词配置文件路径')
    parser.add_argument('--output', default='results', help='输出结果文件名(不含扩展名)')
    parser.add_argument('--chunk-size', type=int, default=1000, help='每批处理的文件数量')
    parser.add_argument('--workers', type=int, default=None, help='最大工作线程数')
    
    args = parser.parse_args()
    
    try:
        processor = FileProcessor(
            config_path=args.config,
            monitor_output=f"{args.output}_processing.csv",
            chunk_size=args.chunk_size,
            max_workers=args.workers
        )
        
        results = []
        if Path(args.path).is_file():
            results = [processor.process_file(args.path)]
        else:
            results = processor.process_directory(args.path)
        
        processor.exporter.export_to_json(results, f"{args.output}.json")
        processor.exporter.export_to_excel(results, f"{args.output}.xlsx")
        
        logger.info(f"处理完成，共处理 {len(results)} 个文件")
    except KeyboardInterrupt:
        print("\n用户中断程序执行")
        sys.exit(1)
    except Exception as e:
        logger.error(f"程序执行出错: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
