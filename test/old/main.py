import os
import yaml
import magic
import time
import logging
import threading
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Any, Iterator, Generator
from concurrent.futures import ThreadPoolExecutor, as_completed
from markitdown import MarkItDown
from dataclasses import dataclass, field
# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('sensitive_detector.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

@dataclass
class ProcessingResult:
    """数据类，用于存储文件处理结果"""
    file_path: str
    mime_type: str
    content: Dict[str, Any]
    sensitive_words: List[Tuple[str, List[int]]] = field(default_factory=list)
    error: Optional[str] = None
    processing_time: float = 0.0
    


class FileTypeDetector:
    """Simplified file type detector using python-magic library"""

    # File extension to MIME type mapping
    MIME_TYPES = {
        # Text files
        ".txt": "text/plain",
        ".csv": "text/csv",
        ".xml": "text/xml",
        ".html": "text/html",
        ".htm": "text/html",
        ".json": "application/json",
        ".yaml": "application/yaml",
        ".yml": "application/yaml",
        ".md": "text/markdown",
        
        # Microsoft Office documents
        ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xls": "application/vnd.ms-excel",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".ppt": "application/vnd.ms-powerpoint",
        ".pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        
        # Compressed files
        ".zip": "application/zip",
        ".rar": "application/x-rar",
        ".7z": "application/x-7z-compressed",
        ".tar": "application/x-tar",
        ".gz": "application/gzip",
        ".bz2": "application/x-bzip2",
        
        # PDF files
        ".pdf": "application/pdf",
        
        # Image files
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
        ".svg": "image/svg+xml",
        ".webp": "image/webp",
        
        # Audio/video files
        ".mp3": "audio/mpeg",
        ".mp4": "video/mp4",
        ".avi": "video/x-msvideo",
        ".mov": "video/quicktime",
        ".wav": "audio/wav",
        
        # Other binary files
        ".bin": "application/octet-stream",
        ".exe": "application/x-msdownload",
        ".dll": "application/x-msdownload",
    }

    # MIME type to extension reverse mapping
    MIME_TO_EXT = {mime: ext for ext, mime in MIME_TYPES.items()}

    def __init__(self):
        """Initialize the file type detector"""
        self.mime = magic.Magic(mime=True)
    
    def get_all_files(self, directory: str) -> List[str]:
        """Get all files in a directory recursively"""
        return [str(f) for f in Path(directory).rglob('*') if f.is_file()]

    def get_file_info(self, file_path: str) -> Dict:
        """
        Get file information including MIME type and file header
        
        Args:
            file_path: Path to the file
            
        Returns:
            Dictionary containing file information
        """
        try:
            # Get MIME type using python-magic
            mime_type = self.mime.from_file(file_path)
            
            # Get file header (first 16 bytes)
            with open(file_path, "rb") as f:
                file_header = f.read(16)
                file_header_hex = file_header.hex().upper()
            
            # Get extension
            file_extension = Path(file_path).suffix.lower()
            
            return {
                "file_path": file_path,
                "mime_type": mime_type,
                "file_extension": file_extension,
                "file_header": file_header_hex,
                "size": os.path.getsize(file_path)
            }
        except Exception as e:
            error_msg = f"Failed to get file info: {e}"
            logger.error(f"{error_msg} - {file_path}")
            return {
                "file_path": file_path,
                "mime_type": "unknown",
                "error": str(e)
            }

    def detect_file_type(self, file_path: str) -> str:
        """
        Detect MIME type of a file
        
        Args:
            file_path: Path to the file
            
        Returns:
            MIME type as string
        """
        try:
            file_info = self.get_file_info(file_path)
            return file_info.get("mime_type", "application/octet-stream")
        except Exception as e:
            logger.error(f"File type detection failed {file_path}: {e}")
            return "application/octet-stream"

    def get_file_header(self, file_path: str, bytes_count: int = 16) -> str:
        """
        Get file header as hex string
        
        Args:
            file_path: Path to the file
            bytes_count: Number of bytes to read (default: 16)
            
        Returns:
            Hex string of file header
        """
        try:
            with open(file_path, 'rb') as f:
                header = f.read(bytes_count)
                return header.hex().upper()
        except Exception as e:
            logger.error(f"Failed to get file header {file_path}: {e}")
            return ""

    def get_extension(self, mime_type: str) -> str:
        """
        Get file extension for a MIME type
        
        Args:
            mime_type: MIME type string
            
        Returns:
            File extension including the dot
        """
        return self.MIME_TO_EXT.get(mime_type, ".unknown")

    def detect_all_files_in_directory(self, directory: str) -> Dict[str, str]:
        """
        Detect MIME types for all files in a directory
        
        Args:
            directory: Directory path
            
        Returns:
            Dictionary mapping file paths to MIME types
        """
        files = self.get_all_files(directory)
        return {file_path: self.detect_file_type(file_path) for file_path in files}


    
    
class ContentExtractor:
    """优化后的文件内容提取器，使用MarkItDown作为主要提取工具"""
    
    def __init__(self):
        """初始化内容提取器"""
        self.md = MarkItDown()  # 启用插件以支持更多格式
        self.fallback_extractors = {
            'application/msword':                                                          self._extract_doc_content,
            'application/vnd.ms-powerpoint':                                               self._extract_ppt_content,
            'application/vnd.ms-excel':                                                    self._extract_docx_content,
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document':     self._extract_docx_content,
            'application/vnd.openxmlformats-officedocument.presentationml.presentation':   self._extract_pptx_content,
            # 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':           self._extract_xlsx_content,
            'text/csv': self._read_csv_file
        }

    def _extract_doc_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from DOC files using win32com"""
        try:
            import win32com.client
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(file_path)
            content = doc.Range().Text
            doc.Close(False)
            word.Quit()
            return {
                'content': content,
                'metadata': {'file_type': 'doc'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'doc'},
                'error': str(e)
            }

    def _extract_docx_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from DOCX files using MarkItDown"""
        try:
            md = MarkItDown(enable_plugins=False)
            result = md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': 'docx'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'docx'},
                'error': str(e)
            }

    def _extract_ppt_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from PPT files using win32com"""
        try:
            md = MarkItDown(enable_plugins=False)
            result = md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': 'ppt'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'ppt'},
                'error': str(e)
            }

    def _extract_xlsx_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from PPTX files using MarkItDown"""
        try:
            md = MarkItDown(enable_plugins=False)
            result = md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': 'pptx'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'pptx'},
                'error': str(e)
            }

    def _extract_pptx_content(self, file_path: str) -> Dict[str, Any]:
        """Extract content from PPTX files using MarkItDown"""
        try:
            md = MarkItDown(enable_plugins=False)
            result = md.convert(file_path)
            return {
                'content': result.text_content,
                'metadata': {'file_type': 'pptx'},
                'error': None
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'pptx'},
                'error': str(e)
            }

    def _is_excel_file(self, file_path: str) -> Tuple[bool, str]:
        """
        检查文件是否为Excel文件，返回(是否是Excel, 具体Excel类型)
        返回的类型可能是: 'xls', 'xlsx', 或 'unknown'
        """
        # 先检查文件头部特征
        try:
            with open(file_path, 'rb') as f:
                header = f.read(2000)
                # XLSX 文件头部特征
                if header.startswith(b'PK'):
                    try:
                        from openpyxl import load_workbook
                        load_workbook(file_path, read_only=True)
                        return True, 'xlsx'
                    except:
                        pass

                # XLS 文件头部特征
                if b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1' in header:
                    try:
                        import xlrd
                        xlrd.open_workbook(file_path)
                        return True, 'xls'
                    except:
                        pass
        except:
            pass

        # 如果头部检查失败，尝试不同的库打开
        try:
            from openpyxl import load_workbook
            load_workbook(file_path, read_only=True)
            return True, 'xlsx'
        except:
            try:
                import xlrd
                xlrd.open_workbook(file_path)
                return True, 'xls'
            except:
                return False, 'unknown'

    def extract_content(self, file_path: str, mime_type: str) -> Dict[str, Any]:
        """提取文件内容，根据文件类型选择合适的提取器"""
        try:
            if not Path(file_path).exists():
                return {'error': f'文件不存在: {file_path}', 'content': '', 'metadata': {}}
                
            # 检查空文件
            if Path(file_path).stat().st_size == 0:
                return {
                    'content': '',
                    'metadata': {'file_type': mime_type},
                    'error': None,
                    'is_empty': True
                }

            # 对于已知的Excel类型，直接使用专用提取器
            if mime_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                try:
                    result = self._read_excel_file(file_path)
                    if not result.get('error'):
                        return result
                except Exception as e:
                    logger.warning(f"Excel提取失败: {str(e)}")

            # 对于其他类型，先检查是否有专用提取器
            if mime_type in self.fallback_extractors:
                try:
                    result = self.fallback_extractors[mime_type](file_path)
                    if not result.get('error'):
                        return result
                except Exception as e:
                    logger.warning(f"专用提取器失败: {str(e)}")

            # 尝试文本文件处理
            try:
                content = self._read_text_file(file_path)
                return {
                    'content': content,
                    'metadata': {'file_type': mime_type},
                    'error': None
                }
            except Exception as text_error:
                pass

            # 最后尝试MarkItDown（避免其错误识别文件类型）
            try:
                result = self.md.convert(file_path)
                return {
                    'content': result.text_content,
                    'metadata': {
                        'file_type': mime_type,
                        'converter': 'markitdown',
                        'additional_metadata': result.metadata if hasattr(result, 'metadata') else {}
                    },
                    'error': None
                }
            except Exception as md_error:
                # 如果所有方法都失败，返回二进制预览
                try:
                    with open(file_path, 'rb') as f:
                        preview = f.read(100).hex()
                    return {
                        'content': f'[Binary content preview: {preview}]',
                        'metadata': {'file_type': mime_type},
                        'error': None,
                        'skipped': True
                    }
                except Exception as bin_error:
                    return {
                        'content': '',
                        'metadata': {'file_type': mime_type},
                        'error': f'Failed to read file: {str(bin_error)}',
                        'skipped': True
                    }

        except Exception as e:
            error_msg = f"提取内容失败 {file_path}: {str(e)}"
            logger.error(error_msg)
            return {
                'content': '',
                'metadata': {'file_type': mime_type},
                'error': error_msg
            }

    def _check_file_type(self, file_path: str) -> bool:
        """检查文件是否为Excel文件"""
        try:
            with open(file_path, 'rb') as f:
                header = f.read(8)
                # Excel 文件的标准头部特征
                excel_signatures = [
                    b'\xD0\xCF\x11\xE0',  # XLS
                    b'PK\x03\x04',        # XLSX
                ]
                return any(header.startswith(sig) for sig in excel_signatures)
        except:
            return False

    def _read_excel_file(self, file_path: str) -> Dict[str, Any]:
        """改进的Excel文件读取方法，增加多重容错"""
        error_messages = []
        file_size = os.path.getsize(file_path)
        
        if not self._check_file_type(file_path):
            return {
                'content': '',
                'metadata': {'file_type': 'unknown'},
                'error': 'Not a valid Excel file',
                'skipped': True
            }
    
        # 检查文件头部特征
        try:
            with open(file_path, 'rb') as f:
                header = f.read(8)
                if not (header.startswith(b'PK') or header.startswith(b'\xD0\xCF\x11\xE0')):
                    return {
                        'content': '',
                        'metadata': {'file_type': 'unknown'},
                        'error': 'Not a valid Excel file format',
                        'skipped': True
                    }
        except Exception as e:
            error_messages.append(f"Header check failed: {str(e)}")

        # XLSX 尝试
        try:
            import pandas as pd
            sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            return self._process_excel_content(sheets)
        except Exception as e:
            error_messages.append(f"openpyxl engine failed: {str(e)}")

        # XLS 尝试
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_path, on_demand=True)
            sheets = {}
            for sheet in workbook.sheets():
                if sheet.nrows > 0:  # 只处理非空sheet
                    data = []
                    for row in range(sheet.nrows):
                        data.append([str(sheet.cell_value(row, col)) for col in range(sheet.ncols)])
                    if data:  # 确保有数据
                        import pandas as pd
                        sheets[sheet.name] = pd.DataFrame(data[1:], columns=data[0] if data else None)
            if sheets:  # 如果成功提取了数据
                return self._process_excel_content(sheets)
        except Exception as e:
            error_messages.append(f"xlrd engine failed: {str(e)}")

        # 如果所有方法都失败，尝试二进制预览
        try:
            with open(file_path, 'rb') as f:
                preview = f.read(1024).hex()
            return {
                'content': f'[Binary content preview: {preview}]',
                'metadata': {'file_type': 'unknown_excel'},
                'error': '\n'.join(error_messages),
                'skipped': True
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'unknown'},
                'error': f'All Excel reading methods failed:\n{"\n".join(error_messages)}',
                'skipped': True
            }

    def _read_text_file(self, file_path: str) -> str:
        """使用多种编码尝试读取文本文件"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
        
        # 首先尝试使用chardet检测编码
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                result = chardet.detect(raw_data)
                if result['confidence'] > 0.8:
                    try:
                        return raw_data.decode(result['encoding'])
                    except:
                        pass
        except ImportError:
            logger.warning("chardet not installed, falling back to manual encoding detection")
        
        # 如果chardet失败或不可用，尝试预定义的编码列表
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
            
        # 如果所有尝试都失败，使用忽略错误的方式读取
        with open(file_path, 'rb') as f:
            return f.read().decode('utf-8', errors='ignore')
            
    def _read_csv_file(self, file_path: str) -> str:
        """改进的CSV文件读取方法"""
        try:
            import pandas as pd
            # 尝试使用pandas读取，自动检测分隔符和编码
            df = pd.read_csv(file_path, sep=None, engine='python')
            return df.to_string(index=False)
        except Exception as pd_error:
            logger.warning(f"Pandas读取CSV失败: {str(pd_error)}")
            # 如果pandas失败，回退到基础CSV读取
            return self._read_text_file(file_path)

    def _extract_with_win32com(self, file_path: str) -> Dict[str, Any]:
        """Word文档提取方法保持不变"""
        # 保持原有的Word文档提取逻辑不变
        pass

    def _read_excel_file(self, file_path: str) -> Dict[str, Any]:
        """改进的Excel文件读取方法，增加多重容错"""
        error_messages = []
        
        # 首先检测Excel类型
        is_excel, excel_type = self._is_excel_file(file_path)
        if is_excel:
            if excel_type == 'xlsx':
                # 优先使用openpyxl方法
                try:
                    import pandas as pd
                    sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
                    return self._process_excel_content(sheets)
                except Exception as e:
                    error_messages.append(f"openpyxl engine failed: {str(e)}")
            elif excel_type == 'xls':
                # 优先使用xlrd方法
                try:
                    import pandas as pd
                    sheets = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
                    return self._process_excel_content(sheets)
                except Exception as e:
                    error_messages.append(f"xlrd engine failed: {str(e)}")
                    
        # 如果特定类型的方法失败，继续尝试其他方法
        
        # 尝试方法1: pandas with openpyxl
        try:
            import pandas as pd
            sheets = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
            return self._process_excel_content(sheets)
        except Exception as e:
            error_messages.append(f"openpyxl engine failed: {str(e)}")
            
        # 尝试方法2: pandas with xlrd
        try:
            import pandas as pd
            sheets = pd.read_excel(file_path, sheet_name=None, engine='xlrd')
            return self._process_excel_content(sheets)
        except Exception as e:
            error_messages.append(f"xlrd engine failed: {str(e)}")
            
        # 尝试方法3: 直接使用xlrd
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            sheets = {}
            for sheet in workbook.sheets():
                data = []
                for row in range(sheet.nrows):
                    data.append([str(sheet.cell_value(row, col)) for col in range(sheet.ncols)])
                sheets[sheet.name] = pd.DataFrame(data[1:], columns=data[0] if data else [])
            return self._process_excel_content(sheets)
        except Exception as e:
            error_messages.append(f"pure xlrd failed: {str(e)}")
            
        # 尝试方法4: 直接使用openpyxl
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                data = []
                for row in ws.rows:
                    data.append([str(cell.value) for cell in row])
                sheets[sheet_name] = pd.DataFrame(data[1:], columns=data[0] if data else [])
            return self._process_excel_content(sheets)
        except Exception as e:
            error_messages.append(f"pure openpyxl failed: {str(e)}")

        # 如果所有方法都失败，尝试作为二进制文件读取预览
        try:
            with open(file_path, 'rb') as f:
                preview = f.read(1024).hex()
            return {
                'content': f'[Binary content preview: {preview}]',
                'metadata': {'file_type': 'unknown_excel'},
                'error': '\n'.join(error_messages),
                'skipped': True
            }
        except Exception as e:
            return {
                'content': '',
                'metadata': {'file_type': 'unknown'},
                'error': f'All Excel reading methods failed:\n{"\n".join(error_messages)}',
                'skipped': True
            }

    def _process_excel_content(self, sheets: Dict) -> Dict[str, Any]:
        """处理Excel内容的辅助方法"""
        content_parts = []
        metadata = {
            'sheets': list(sheets.keys()),
            'total_sheets': len(sheets),
            'row_counts': {}
        }
        
        for sheet_name, df in sheets.items():
            content_parts.append(f"\n=== Sheet: {sheet_name} ===\n")
            content_parts.append(df.to_string(index=False))
            metadata['row_counts'][sheet_name] = len(df)
        
        return {
            'content': '\n'.join(content_parts),
            'metadata': {
                'file_type': 'excel',
                **metadata
            },
            'error': None
        }
    
    
class SensitiveChecker:
    """敏感内容检查器"""
    
    def __init__(self, config_path: str = "sensitive_config.yaml"):
        """初始化敏感词配置"""
        self.config = self._load_config(config_path)
        
    def _load_config(self, config_path: str) -> Dict:
        """加载敏感词配置文件"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except Exception as e:
            logger.error(f"加载敏感词配置失败: {e}")
            return {}
            
    def check_content(self, text: str) -> List[Tuple[str, List[int]]]:
        """
        检查文本中的敏感词
        返回: [(敏感词, [出现位置列表]), ...]
        """
        results = []
        
        # 检查安全标记
        for mark in self.config.get('security_marks', []):
            positions = []
            start = 0
            while True:
                pos = text.find(mark, start)
                if pos == -1:
                    break
                positions.append(pos)
                start = pos + 1
            if positions:
                results.append((mark, positions))
        
        # 检查分类敏感词
        for category, category_config in self.config.get('sensitive_patterns', {}).items():
            for keyword in category_config.get('keywords', []):
                positions = []
                start = 0
                while True:
                    pos = text.find(keyword, start)
                    if pos == -1:
                        break
                    positions.append(pos)
                    start = pos + 1
                if positions:
                    results.append((keyword, positions))
        
        # 检查结构化模式
        for pattern, weight in self.config.get('structured_patterns', {}).items():
            import re
            matches = list(re.finditer(pattern, text))
            if matches:
                positions = [m.start() for m in matches]
                results.append((pattern, positions))
                
        # 检查数字模式
        for pattern in self.config.get('number_patterns', []):
            import re
            matches = list(re.finditer(pattern, text))
            if matches:
                positions = [m.start() for m in matches]
                results.append((pattern, positions))
        
        return results
    

class ResultExporter:
    """处理结果导出器"""
    
    def export_to_json(self, results: List[ProcessingResult], output_path: str):
        """导出结果到JSON文件"""
        import json
        
        export_data = []
        for result in results:
            export_data.append({
                'file_path': result.file_path,
                'mime_type': result.mime_type,
                'content_preview': result.content.get('content', '')[:200],
                'sensitive_words': [
                    {'word': word, 'positions': positions}
                    for word, positions in result.sensitive_words
                ],
                'error': result.error,
                'processing_time': result.processing_time
            })
            
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
    
    def export_to_excel(self, results: List[ProcessingResult], output_path: str):
        """导出结果到Excel文件"""
        import pandas as pd
        
        data = []
        for result in results:
            sensitive_words_str = '; '.join([
                f"{word}({len(positions)}次)" 
                for word, positions in result.sensitive_words
            ])
            
            data.append({
                '文件路径': result.file_path,
                '文件类型': result.mime_type,
                '敏感词统计': sensitive_words_str,
                '处理时间(秒)': round(result.processing_time, 3),
                '错误信息': result.error or ''
            })
            
        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        
        
        
class ResultMonitor:
    """结果监控器，用于实时输出处理进度和结果"""
    
    def __init__(self, output_csv: str = "processing_results.csv"):
        self.output_csv = output_csv
        self._init_csv()
        
    def _init_csv(self):
        """初始化CSV文件"""
        import csv
        with open(self.output_csv, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                '处理时间',
                '文件路径',
                '文件类型',
                '文件大小(bytes)',
                '是否为空',
                '敏感词数量',
                '敏感词列表',
                '处理时长(秒)',
                '状态',
                '错误信息'
            ])
    
    def record_result(self, result: ProcessingResult):
        """记录单个处理结果"""
        import csv
        from datetime import datetime
        
        # 获取文件大小
        try:
            file_size = Path(result.file_path).stat().st_size
        except:
            file_size = 0
            
        # 统计敏感词
        sensitive_words_count = len(result.sensitive_words)
        sensitive_words_list = '; '.join([
            f"{word}({len(positions)}处)" 
            for word, positions in result.sensitive_words
        ])
        
        # 确定状态
        if result.error:
            status = '失败'
        elif result.content.get('is_empty'):
            status = '空文件'
        elif result.content.get('skipped'):
            status = '已跳过'
        else:
            status = '成功'
            
        # 写入CSV
        with open(self.output_csv, 'a', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                result.file_path,
                result.mime_type,
                file_size,
                'Yes' if result.content.get('is_empty') else 'No',
                sensitive_words_count,
                sensitive_words_list,
                round(result.processing_time, 3),
                status,
                result.error or ''
            ])
        
        # 打印实时进度
        print(f"[{datetime.now().strftime('%H:%M:%S')}] "
              f"处理文件: {result.file_path}")
        print(f"  类型: {result.mime_type}")
        print(f"  大小: {file_size} bytes")
        print(f"  状态: {status}")
        if sensitive_words_count > 0:
            print(f"  发现敏感词: {sensitive_words_list}")
        if result.error:
            print(f"  错误: {result.error}")
        print("-" * 80)
        
        
        
class FileProcessor:
    """优化后的文件处理器主类"""
    
    def __init__(self, config_path: str = "sensitive_config.yaml", 
                 monitor_output: str = "processing_results.csv",
                 chunk_size: int = 1000,
                 max_workers: Optional[int] = None):  # 新增参数
        self.detector = FileTypeDetector()
        self.extractor = ContentExtractor()
        self.checker = SensitiveChecker(config_path)
        self.exporter = ResultExporter()
        self.monitor = ResultMonitor(monitor_output)
        self.chunk_size = chunk_size
        self.max_workers = max_workers or (os.cpu_count() or 1) * 2  # 新增
        self._mime_cache = {}
        self._file_size_cache = {}
        
    def _scan_directory(self, directory: str) -> Iterator[List[str]]:
        """使用生成器分批扫描目录文件"""
        current_chunk = []
        
        for entry in os.scandir(directory):
            try:
                if entry.is_file(follow_symlinks=False):
                    current_chunk.append(entry.path)
                    if len(current_chunk) >= self.chunk_size:
                        yield current_chunk
                        current_chunk = []
                elif entry.is_dir(follow_symlinks=False):
                    # 递归处理子目录
                    sub_path = entry.path
                    for sub_chunk in self._scan_directory(sub_path):
                        yield sub_chunk
            except (PermissionError, OSError) as e:
                logger.warning(f"访问文件/目录出错 {entry.path}: {e}")
                continue
                    
        if current_chunk:
            yield current_chunk
            
    def _preload_file_info(self, file_paths: List[str]):
        """预加载文件信息以减少IO操作"""
        for file_path in file_paths:
            try:
                # 批量预加载文件大小
                stat = os.stat(file_path)
                self._file_size_cache[file_path] = stat.st_size
                
                # 批量预加载MIME类型
                mime_type = self.detector.detect_file_type(file_path)
                self._mime_cache[file_path] = mime_type
            except Exception as e:
                logger.warning(f"预加载文件信息失败 {file_path}: {e}")
                
    def _process_file_batch(self, file_paths: List[str]) -> List[ProcessingResult]:
        """处理一批文件"""
        results = []
        
        # 首先预加载这批文件的信息
        self._preload_file_info(file_paths)
        
        # 使用线程池处理这批文件
        with ThreadPoolExecutor(max_workers=min(len(file_paths), self.max_workers)) as executor:
            future_to_file = {
                executor.submit(self.process_file, file_path): file_path 
                for file_path in file_paths
            }
            
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    results.append(result)
                    self.monitor.record_result(result)
                except Exception as e:
                    logger.error(f"处理文件异常 {file_path}: {e}")
                    
        return results
        

    def process_directory(self, directory: str) -> List[ProcessingResult]:
        """处理目录下的所有文件"""
        results = []
        total_files = sum(1 for _ in Path(directory).rglob('*') if _.is_file())
        completed = 0
        
        print(f"\n开始处理目录: {directory}")
        print(f"共发现 {total_files} 个文件")
        print("=" * 80)
        
        try:
            # 分批处理文件
            for file_chunk in self._scan_directory(directory):
                chunk_results = self._process_file_batch(file_chunk)
                results.extend(chunk_results)
                
                completed += len(file_chunk)
                print(f"\n已完成: {completed}/{total_files} ({completed/total_files*100:.1f}%)")
                
                # 定期清理缓存
                if completed % (self.chunk_size * 10) == 0:
                    self._mime_cache.clear()
                    self._file_size_cache.clear()
                    
        except KeyboardInterrupt:
            print("\n用户中断处理...")
        except Exception as e:
            logger.error(f"处理目录时发生错误: {e}")
        finally:
            # 导出已处理的结果
            if results:
                try:
                    self.exporter.export_to_json(results, "interrupted_results.json")
                    self.exporter.export_to_excel(results, "interrupted_results.xlsx")
                except Exception as e:
                    logger.error(f"导出中断结果失败: {e}")
        
        print(f"\n处理完成！共处理 {len(results)}/{total_files} 个文件")
        print(f"详细结果已保存至: {self.monitor.output_csv}")
        print("=" * 80)
        
        return results
        
    def process_file(self, file_path: str) -> ProcessingResult:
        """处理单个文件(优化版本)"""
        start_time = time.time()
        
        try:
            # 使用缓存的MIME类型
            mime_type = self._mime_cache.get(file_path)
            if not mime_type:
                mime_type = self.detector.detect_file_type(file_path)
                
            logger.info(f"Processing file: {file_path} (type: {mime_type})")
            
            # 使用缓存的文件大小
            file_size = self._file_size_cache.get(file_path, 0)
            if file_size == 0:
                # 如果文件为空，直接返回结果
                return ProcessingResult(
                    file_path=file_path,
                    mime_type=mime_type,
                    content={'is_empty': True},
                    sensitive_words=[],
                    error=None,
                    processing_time=time.time() - start_time
                )
            
            # 提取内容
            content = self.extractor.extract_content(file_path, mime_type)
            
            # 检查处理结果
            if content.get('skipped') or content.get('error'):
                return ProcessingResult(
                    file_path=file_path,
                    mime_type=mime_type,
                    content=content,
                    sensitive_words=[],
                    error=content.get('error'),
                    processing_time=time.time() - start_time
                )
            
            # 检查敏感词
            sensitive_words = self.checker.check_content(content.get('content', ''))
            if sensitive_words:
                logger.info(f"Found {len(sensitive_words)} sensitive words in {file_path}")
            
            return ProcessingResult(
                file_path=file_path,
                mime_type=mime_type,
                content=content,
                sensitive_words=sensitive_words,
                error=None,
                processing_time=time.time() - start_time
            )
            
        except Exception as e:
            error_msg = f"处理文件失败: {str(e)}"
            logger.error(f"{error_msg} - {file_path}")
            return ProcessingResult(
                file_path=file_path,
                mime_type="unknown",
                content={'error': error_msg},
                processing_time=time.time() - start_time
            )


'''
class ResultMonitor:
    """Result monitor with simplified CSV output"""
    
    def __init__(self, output_csv: str = "processing_results.csv"):
        self.output_csv = output_csv
        self._init_csv()
        
    def _init_csv(self):
        """Initialize CSV file with simplified headers"""
        import csv
        with open(self.output_csv, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                '文件名',
                '是否敏感'
            ])
    
    def record_result(self, result: ProcessingResult):
        """Record processing result with simplified output"""
        import csv
        from pathlib import Path
        
        # Get just the filename
        filename = Path(result.file_path).name
        
        # Determine if file contains sensitive content (0 or 1)
        is_sensitive = 1 if result.sensitive_words else 0
            
        # Write to CSV
        with open(self.output_csv, 'a', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                filename,
                is_sensitive
            ])
        
        # Print real-time progress (keeping this for monitoring purposes)
        print(f"处理文件: {filename}")
        print(f"是否敏感: {'是' if is_sensitive else '否'}")
        if result.error:
            print(f"错误: {result.error}")
        print("-" * 40)
'''
        
        
        
        
        
class FileProcessor:
    """优化后的文件处理器主类"""
    
    def __init__(self, config_path: str = "sensitive_config.yaml", 
                 monitor_output: str = "processing_results.csv",
                 chunk_size: int = 1000,
                 max_workers: Optional[int] = None):  # 新增参数
        self.detector = FileTypeDetector()
        self.extractor = ContentExtractor()
        self.checker = SensitiveChecker(config_path)
        self.exporter = ResultExporter()
        self.monitor = ResultMonitor(monitor_output)
        self.chunk_size = chunk_size
        self.max_workers = max_workers or (os.cpu_count() or 1) * 2  # 新增
        self._mime_cache = {}
        self._file_size_cache = {}
        
    def _scan_directory(self, directory: str) -> Iterator[List[str]]:
        """使用生成器分批扫描目录文件"""
        current_chunk = []
        
        for entry in os.scandir(directory):
            try:
                if entry.is_file(follow_symlinks=False):
                    current_chunk.append(entry.path)
                    if len(current_chunk) >= self.chunk_size:
                        yield current_chunk
                        current_chunk = []
                elif entry.is_dir(follow_symlinks=False):
                    # 递归处理子目录
                    sub_path = entry.path
                    for sub_chunk in self._scan_directory(sub_path):
                        yield sub_chunk
            except (PermissionError, OSError) as e:
                logger.warning(f"访问文件/目录出错 {entry.path}: {e}")
                continue
                    
        if current_chunk:
            yield current_chunk
            
    def _preload_file_info(self, file_paths: List[str]):
        """预加载文件信息以减少IO操作"""
        for file_path in file_paths:
            try:
                # 批量预加载文件大小
                stat = os.stat(file_path)
                self._file_size_cache[file_path] = stat.st_size
                
                # 批量预加载MIME类型
                mime_type = self.detector.detect_file_type(file_path)
                self._mime_cache[file_path] = mime_type
            except Exception as e:
                logger.warning(f"预加载文件信息失败 {file_path}: {e}")
                
    def _process_file_batch(self, file_paths: List[str]) -> List[ProcessingResult]:
        """处理一批文件"""
        results = []
        
        # 首先预加载这批文件的信息
        self._preload_file_info(file_paths)
        
        # 使用线程池处理这批文件
        with ThreadPoolExecutor(max_workers=min(len(file_paths), self.max_workers)) as executor:
            future_to_file = {
                executor.submit(self.process_file, file_path): file_path 
                for file_path in file_paths
            }
            
            for future in as_completed(future_to_file):
                file_path = future_to_file[future]
                try:
                    result = future.result()
                    results.append(result)
                    self.monitor.record_result(result)
                except Exception as e:
                    logger.error(f"处理文件异常 {file_path}: {e}")
                    
        return results
        

    def process_directory(self, directory: str) -> List[ProcessingResult]:
        """处理目录下的所有文件"""
        results = []
        total_files = sum(1 for _ in Path(directory).rglob('*') if _.is_file())
        completed = 0
        
        print(f"\n开始处理目录: {directory}")
        print(f"共发现 {total_files} 个文件")
        print("=" * 80)
        
        try:
            # 分批处理文件
            for file_chunk in self._scan_directory(directory):
                chunk_results = self._process_file_batch(file_chunk)
                results.extend(chunk_results)
                
                completed += len(file_chunk)
                print(f"\n已完成: {completed}/{total_files} ({completed/total_files*100:.1f}%)")
                
                # 定期清理缓存
                if completed % (self.chunk_size * 10) == 0:
                    self._mime_cache.clear()
                    self._file_size_cache.clear()
                    
        except KeyboardInterrupt:
            print("\n用户中断处理...")
        except Exception as e:
            logger.error(f"处理目录时发生错误: {e}")
        finally:
            # 导出已处理的结果
            if results:
                try:
                    self.exporter.export_to_json(results, "interrupted_results.json")
                    self.exporter.export_to_excel(results, "interrupted_results.xlsx")
                except Exception as e:
                    logger.error(f"导出中断结果失败: {e}")
        
        print(f"\n处理完成！共处理 {len(results)}/{total_files} 个文件")
        print(f"详细结果已保存至: {self.monitor.output_csv}")
        print("=" * 80)
        
        return results
        
    def process_file(self, file_path: str) -> ProcessingResult:
        """处理单个文件(优化版本)"""
        start_time = time.time()
        
        try:
            # 使用缓存的MIME类型
            mime_type = self._mime_cache.get(file_path)
            if not mime_type:
                mime_type = self.detector.detect_file_type(file_path)
                
            logger.info(f"Processing file: {file_path} (type: {mime_type})")
            
            # 使用缓存的文件大小
            file_size = self._file_size_cache.get(file_path, 0)
            if file_size == 0:
                # 如果文件为空，直接返回结果
                return ProcessingResult(
                    file_path=file_path,
                    mime_type=mime_type,
                    content={'is_empty': True},
                    sensitive_words=[],
                    error=None,
                    processing_time=time.time() - start_time
                )
            
            # 提取内容
            content = self.extractor.extract_content(file_path, mime_type)
            
            # 检查处理结果
            if content.get('skipped') or content.get('error'):
                return ProcessingResult(
                    file_path=file_path,
                    mime_type=mime_type,
                    content=content,
                    sensitive_words=[],
                    error=content.get('error'),
                    processing_time=time.time() - start_time
                )
            
            # 检查敏感词
            sensitive_words = self.checker.check_content(content.get('content', ''))
            if sensitive_words:
                logger.info(f"Found {len(sensitive_words)} sensitive words in {file_path}")
            
            return ProcessingResult(
                file_path=file_path,
                mime_type=mime_type,
                content=content,
                sensitive_words=sensitive_words,
                error=None,
                processing_time=time.time() - start_time
            )
            
        except Exception as e:
            error_msg = f"处理文件失败: {str(e)}"
            logger.error(f"{error_msg} - {file_path}")
            return ProcessingResult(
                file_path=file_path,
                mime_type="unknown",
                content={'error': error_msg},
                processing_time=time.time() - start_time
            )



def main():
    """主函数"""
    import argparse
    import sys  # 新增
    
    parser = argparse.ArgumentParser(description='文件敏感内容检测工具')
    parser.add_argument('path', help='要处理的文件或目录路径')
    parser.add_argument('--config', default='sensitive_config.yaml', help='敏感词配置文件路径')
    parser.add_argument('--output', default='results', help='输出结果文件名(不含扩展名)')
    parser.add_argument('--chunk-size', type=int, default=1000, help='每批处理的文件数量')
    parser.add_argument('--workers', type=int, default=None, help='最大工作线程数')  # 新增
    
    args = parser.parse_args()
    
    try:
        # 初始化处理器
        processor = FileProcessor(
            config_path=args.config,
            monitor_output=f"{args.output}_processing.csv",
            chunk_size=args.chunk_size,
            max_workers=args.workers  # 新增
        )
        
        # 处理文件或目录
        results = []
        if Path(args.path).is_file():
            results = [processor.process_file(args.path)]
        else:
            results = processor.process_directory(args.path)
        
        # 导出结果
        processor.exporter.export_to_json(results, f"{args.output}.json")
        processor.exporter.export_to_excel(results, f"{args.output}.xlsx")
        
        logger.info(f"处理完成，共处理 {len(results)} 个文件")
        
    except KeyboardInterrupt:
        print("\n用户中断程序执行")
        sys.exit(1)
    except Exception as e:
        logger.error(f"程序执行出错: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
